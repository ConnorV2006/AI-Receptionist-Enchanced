import os
import csv
from io import StringIO, BytesIO
from datetime import datetime, timedelta

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    session, Response, send_file
)
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func
import openpyxl

# -------------------------------------------------
# App Setup
# -------------------------------------------------
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev_secret")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# -------------------------------------------------
# Models
# -------------------------------------------------
class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    is_superadmin = db.Column(db.Boolean, default=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Clinic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    slug = db.Column(db.String(80), unique=True, nullable=False)
    logo_url = db.Column(db.String(200))
    primary_color = db.Column(db.String(20), default="#1f6feb")


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    admin_id = db.Column(db.Integer, db.ForeignKey("admin.id"))
    action = db.Column(db.String(100), nullable=False)  # login, logout, clock_in, clock_out
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    details = db.Column(db.Text)
    admin = db.relationship("Admin", backref="audit_logs")


class Shift(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    admin_id = db.Column(db.Integer, db.ForeignKey("admin.id"))
    clock_in = db.Column(db.DateTime, default=datetime.utcnow)
    clock_out = db.Column(db.DateTime, nullable=True)
    admin = db.relationship("Admin", backref="shifts")

    @property
    def duration_hours(self):
        if self.clock_in and self.clock_out:
            return round((self.clock_out - self.clock_in).total_seconds() / 3600, 2)
        return None

# -------------------------------------------------
# Helpers
# -------------------------------------------------
def get_current_admin():
    admin_id = session.get("admin_id")
    return Admin.query.get(admin_id) if admin_id else None

def log_action(admin_id, action, details=""):
    db.session.add(AuditLog(admin_id=admin_id, action=action, details=details))
    db.session.commit()

def build_weekly_summary_rows():
    """Return list of (username, total_hours_last_7_days)."""
    week_start = datetime.utcnow() - timedelta(days=7)
    rows = (
        db.session.query(
            Admin.username,
            func.sum(func.extract("epoch", Shift.clock_out - Shift.clock_in) / 3600)
        )
        .join(Shift.admin)
        .filter(Shift.clock_in >= week_start, Shift.clock_out != None)
        .group_by(Admin.username)
        .all()
    )
    return rows

def create_payroll_workbook(shifts):
    """
    Build an Excel workbook in memory with:
    - Shifts sheet (detailed)
    - Payroll Summary (last 4 weeks, overtime flag > 40)
    - Monthly Summary
    Returns BytesIO pointing at file start.
    """
    wb = openpyxl.Workbook()

    # Sheet 1: Detailed Shifts
    ws1 = wb.active
    ws1.title = "Shifts"
    ws1.append(["Staff", "Clock In", "Clock Out", "Hours"])
    for s in shifts:
        ws1.append([
            s.admin.username,
            s.clock_in.strftime("%Y-%m-%d %H:%M"),
            s.clock_out.strftime("%Y-%m-%d %H:%M") if s.clock_out else "Active",
            s.duration_hours if s.duration_hours else "-"
        ])

    # Sheet 2: Payroll Summary (last 4 weeks)
    ws2 = wb.create_sheet("Payroll Summary")
    ws2.append(["Staff", "Week Start", "Week End", "Total Hours", "Overtime?"])
    today = datetime.utcnow().date()
    for staff in Admin.query.all():
        for w in range(4):
            week_end = today - timedelta(days=(7 * w))
            week_start = week_end - timedelta(days=6)
            staff_shifts = Shift.query.filter(
                Shift.admin_id == staff.id,
                Shift.clock_in >= week_start,
                Shift.clock_in <= week_end,
                Shift.clock_out != None
            ).all()
            total_hours = sum(s.duration_hours or 0 for s in staff_shifts)
            overtime = "YES" if total_hours > 40 else "NO"
            ws2.append([
                staff.username,
                week_start.strftime("%Y-%m-%d"),
                week_end.strftime("%Y-%m-%d"),
                round(total_hours, 2),
                overtime
            ])

    # Sheet 3: Monthly Summary
    ws3 = wb.create_sheet("Monthly Summary")
    ws3.append(["Staff", "Month", "Total Hours"])
    for staff in Admin.query.all():
        staff_shifts = Shift.query.filter(
            Shift.admin_id == staff.id,
            Shift.clock_out != None
        ).all()
        monthly_totals = {}
        for s in staff_shifts:
            month_key = s.clock_in.strftime("%Y-%m")
            monthly_totals[month_key] = monthly_totals.get(month_key, 0) + (s.duration_hours or 0)
        for month, total in monthly_totals.items():
            ws3.append([staff.username, month, round(total, 2)])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# -------------------------------------------------
# Routes
# -------------------------------------------------
@app.route("/")
def dashboard():
    admin = get_current_admin()
    return render_template("dashboard.html", admin=admin)

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        admin = Admin.query.filter_by(username=username).first()
        if admin and admin.check_password(password):
            session["admin_id"] = admin.id
            flash("Login successful!", "success")
            log_action(admin.id, "login", f"{username} logged in.")
            return redirect(url_for("dashboard"))
        flash("Invalid username or password", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    admin = get_current_admin()
    if admin:
        log_action(admin.id, "logout", f"{admin.username} logged out.")
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))

@app.route("/clock-in")
def clock_in():
    admin = get_current_admin()
    if not admin:
        flash("You must be logged in to clock in.", "error")
        return redirect(url_for("login"))
    active = Shift.query.filter_by(admin_id=admin.id, clock_out=None).first()
    if active:
        flash("You are already clocked in!", "error")
    else:
        db.session.add(Shift(admin_id=admin.id))
        db.session.commit()
        log_action(admin.id, "clock_in", "Staff clocked in.")
        flash("Clock-in successful!", "success")
    return redirect(url_for("dashboard"))

@app.route("/clock-out")
def clock_out():
    admin = get_current_admin()
    if not admin:
        flash("You must be logged in to clock out.", "error")
        return redirect(url_for("login"))
    active = Shift.query.filter_by(admin_id=admin.id, clock_out=None).first()
    if not active:
        flash("You are not clocked in!", "error")
    else:
        active.clock_out = datetime.utcnow()
        db.session.commit()
        log_action(admin.id, "clock_out", "Staff clocked out.")
        flash("Clock-out successful!", "success")
    return redirect(url_for("dashboard"))

@app.route("/reports", methods=["GET", "POST"])
def reports():
    admin = get_current_admin()
    if not admin or not admin.is_superadmin:
        flash("Permission denied.", "error")
        return redirect(url_for("dashboard"))

    # Filters
    staff_id = request.form.get("staff_id")
    date_from = request.form.get("date_from")
    date_to = request.form.get("date_to")

    query = Shift.query
    if staff_id and staff_id != "all":
        query = query.filter_by(admin_id=int(staff_id))
    if date_from:
        query = query.filter(Shift.clock_in >= date_from)
    if date_to:
        query = query.filter(Shift.clock_in <= date_to)

    shifts = query.order_by(Shift.clock_in.desc()).all()
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(50).all()
    weekly_hours = build_weekly_summary_rows()

    return render_template(
        "reports.html",
        admin=admin,
        shifts=shifts,
        logs=logs,
        weekly_hours=weekly_hours,
        all_staff=Admin.query.all()
    )

# -------------------------
# Export CSV
# -------------------------
@app.route("/reports/export/csv")
def export_csv():
    admin = get_current_admin()
    if not admin or not admin.is_superadmin:
        flash("Permission denied.", "error")
        return redirect(url_for("dashboard"))

    shifts = Shift.query.order_by(Shift.clock_in.desc()).all()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Staff", "Clock In", "Clock Out", "Hours"])
    for s in shifts:
        writer.writerow([
            s.admin.username,
            s.clock_in.strftime("%Y-%m-%d %H:%M"),
            s.clock_out.strftime("%Y-%m-%d %H:%M") if s.clock_out else "Active",
            s.duration_hours if s.duration_hours else "-"
        ])
    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = "attachment; filename=shifts.csv"
    return resp

# -------------------------
# Export Excel (Payroll)
# -------------------------
@app.route("/reports/export/excel")
def export_excel():
    admin = get_current_admin()
    if not admin or not admin.is_superadmin:
        flash("Permission denied.", "error")
        return redirect(url_for("dashboard"))

    shifts = Shift.query.order_by(Shift.clock_in.desc()).all()
    buf = create_payroll_workbook(shifts)
    return send_file(
        buf, as_attachment=True, download_name="shifts_with_payroll.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------------------------------------
# Run (local only)
# -------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True)
