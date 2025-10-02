import os
from datetime import datetime, timedelta
from io import BytesIO

from flask import Flask
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

# Import your app + models
from app import app, db, Admin, Shift

# ================
# Generate Payroll Excel
# ================
def build_excel_report():
    wb = openpyxl.Workbook()

    # Sheet 1: Detailed Shifts
    ws1 = wb.active
    ws1.title = "Shifts"
    ws1.append(["Staff", "Clock In", "Clock Out", "Hours"])

    shifts = Shift.query.order_by(Shift.clock_in.desc()).all()
    for shift in shifts:
        ws1.append([
            shift.admin.username,
            shift.clock_in.strftime("%Y-%m-%d %H:%M"),
            shift.clock_out.strftime("%Y-%m-%d %H:%M") if shift.clock_out else "Active",
            shift.duration_hours if shift.duration_hours else "-"
        ])

    # Sheet 2: Weekly Payroll Summary
    ws2 = wb.create_sheet("Weekly Payroll")
    ws2.append(["Staff", "Week Start", "Week End", "Total Hours", "Overtime?"])

    today = datetime.utcnow().date()
    for staff in Admin.query.all():
        for week in range(4):
            week_end = today - timedelta(days=(7 * week))
            week_start = week_end - timedelta(days=6)

            staff_shifts = Shift.query.filter(
                Shift.admin_id == staff.id,
                Shift.clock_in >= week_start,
                Shift.clock_in <= week_end,
                Shift.clock_out != None
            ).all()

            total_hours = sum(s.duration_hours or 0 for s in staff_shifts)
            overtime = "YES" if total_hours > 40 else "NO"

            ws2.append([
                staff.username,
                week_start.strftime("%Y-%m-%d"),
                week_end.strftime("%Y-%m-%d"),
                round(total_hours, 2),
                overtime
            ])

    # Sheet 3: Monthly Summary
    ws3 = wb.create_sheet("Monthly Summary")
    ws3.append(["Staff", "Month", "Total Hours"])

    for staff in Admin.query.all():
        monthly_totals = {}
        for s in Shift.query.filter_by(admin_id=staff.id).filter(Shift.clock_out != None).all():
            key = s.clock_in.strftime("%Y-%m")
            monthly_totals[key] = monthly_totals.get(key, 0) + (s.duration_hours or 0)

        for month, total in monthly_totals.items():
            ws3.append([staff.username, month, round(total, 2)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ================
# Send Email via SendGrid
# ================
def send_report_via_email():
    output = build_excel_report()

    # Find boss email
    superadmin = Admin.query.filter_by(is_superadmin=True).first()
    if not superadmin or not superadmin.username:
        print("⚠️ No superadmin email found")
        return

    boss_email = os.environ.get("BOSS_EMAIL", "admin@example.com")
    sg_api_key = os.environ.get("SENDGRID_API_KEY")

    if not sg_api_key:
        print("⚠️ Missing SENDGRID_API_KEY in env")
        return

    message = Mail(
        from_email="noreply@ai-receptionist.com",
        to_emails=boss_email,
        subject="📊 Weekly Payroll Report",
        plain_text_content="Attached is the latest payroll Excel report."
    )

    # Attach Excel file
    import base64
    file_content = base64.b64encode(output.read()).decode()
    message.add_attachment(file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "payroll_report.xlsx", "attachment")

    try:
        sg = SendGridAPIClient(sg_api_key)
        response = sg.send(message)
        print(f"✅ Payroll report sent! Status {response.status_code}")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")

if __name__ == "__main__":
    with app.app_context():
        send_report_via_email()
